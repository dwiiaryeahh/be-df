"""
Export service - Handle PDF dan Excel export untuk crawling data
"""
from typing import Dict
from sqlalchemy.orm import Session
from datetime import datetime
from app.db.models import Campaign, Crawling
import io, os

# For PDF
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors

# For Excel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def get_campaign_with_crawling(db: Session, campaign_id: int) -> Dict:
    """Get campaign dan crawling data"""
    campaign = db.query(Campaign).filter(Campaign.id == campaign_id).first()
    
    if not campaign:
        return {"status": "error", "message": "Campaign not found"}
    
    crawlings = db.query(Crawling).filter(
        Crawling.campaign_id == campaign_id
    ).all()
    
    return {
        "status": "success",
        "campaign": campaign,
        "crawlings": crawlings
    }


def generate_pdf(db: Session, campaign_id: int) -> bytes:
    """Generate PDF using reportlab with specific layout"""
    data = get_campaign_with_crawling(db, campaign_id)
    
    if data["status"] == "error":
        return None
    
    campaign = data["campaign"]
    crawlings = data["crawlings"]
    
    target_map = {}
    if campaign.target_info:
        for t in campaign.target_info:
            if isinstance(t, dict) and 'imsi' in t:
                target_map[t['imsi']] = t
    
    imsi_detected = len(crawlings)
    alert_count = 0
    
    table_data = [['No', 'IMSI', 'Time', 'Count', 'Alert Status', 'Alert Name']]
    
    for idx, crawl in enumerate(crawlings, 1):
        if hasattr(crawl, 'timestamp') and crawl.timestamp:
            if isinstance(crawl.timestamp, str):
                timestamp_str = crawl.timestamp
            else:
                timestamp_str = crawl.timestamp.strftime("%Y-%m-%d %H:%M:%S")
        else:
            timestamp_str = "-"
            
        tgt = target_map.get(crawl.imsi)
        alert_status = tgt.get('alert_status') if tgt and tgt.get('alert_status') else "-"
        alert_name = tgt.get('name') if tgt and tgt.get('name') else "-"
        
        if alert_status != "-":
            alert_count += 1
            
        count_val = str(crawl.count) if crawl.count is not None else "0"
        
        table_data.append([
            str(idx),
            crawl.imsi,
            timestamp_str,
            count_val,
            alert_status,
            alert_name
        ])
        
    start_date_str = campaign.start_scan.strftime("%Y-%m-%d (%H:%M:%S)") if campaign.start_scan else "-"
    end_date_str = campaign.stop_scan.strftime("%Y-%m-%d (%H:%M:%S)") if campaign.stop_scan else "-"
    
    duration_str = "-"
    if campaign.start_scan and campaign.stop_scan:
        delta = campaign.stop_scan - campaign.start_scan
        total_seconds = int(delta.total_seconds())
        minutes = total_seconds // 60
        seconds = total_seconds % 60
        duration_str = f"{minutes} Minutes {seconds} Seconds"
    elif campaign.duration:
        pass

    pdf_buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        pdf_buffer, 
        pagesize=letter,
        leftMargin=0.5*inch,
        rightMargin=0.5*inch,
        topMargin=0.8*inch,
        bottomMargin=0.8*inch
    )
    
    def draw_header_footer(canvas, doc):
        canvas.saveState()
        
        logo_path = os.getenv('APP_LOGO', 'app/asset/logo.png')
        app_name = os.getenv("APP_NAME", "Backpack DF")
        app_font = os.getenv("APP_FONT", "Helvetica")
        
        if os.path.exists(logo_path):
            # Raised y from 10.3 to 10.35
            canvas.drawImage(logo_path, 0.5*inch, 10.35*inch, width=0.4*inch, height=0.4*inch, preserveAspectRatio=True, mask='auto')
            
            # App Name next to Logo
            # Raised y from 10.45 to 10.48
            try:
                canvas.setFont(f'{app_font}-Bold', 14)
            except:
                canvas.setFont('Helvetica-Bold', 14)
                
            canvas.setFillColor(colors.HexColor('#2ecc71'))
            canvas.drawString(1.0*inch, 10.48*inch, app_name)
        else:
            try:
                canvas.setFont(f'{app_font}-Bold', 14)
            except:
                canvas.setFont('Helvetica-Bold', 14)
                
            canvas.setFillColor(colors.HexColor('#2ecc71'))
            canvas.drawString(0.5*inch, 10.48*inch, app_name)
            
        # Exported Date - Thinner/Lighter
        exported_str = f"Exported: {datetime.now().strftime('%d/%m/%Y %H:%M')} WIB"
        canvas.setFont('Helvetica', 8)
        canvas.setFillColor(colors.darkgrey)
        canvas.drawRightString(8.0*inch, 10.48*inch, exported_str)
        
        canvas.setLineWidth(0.5)
        canvas.setStrokeColor(colors.grey)
        canvas.line(0.5*inch, 0.75*inch, 8.0*inch, 0.75*inch)
        
        canvas.setFont('Helvetica', 9)
        canvas.setFillColor(colors.grey)
        canvas.drawString(0.5*inch, 0.5*inch, f"{app_name} - Report")
        canvas.drawRightString(8.0*inch, 0.5*inch, f"Page {doc.page}")
        
        canvas.restoreState()

    elements = []
    styles = getSampleStyleSheet()
    
    elements.append(Paragraph(campaign.name, ParagraphStyle(
        'CampaignTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=6,
        textColor=colors.black
    )))
    
    mode_text = f"Mode: {campaign.mode}" if campaign.mode else "Mode: -"
    elements.append(Paragraph(mode_text, ParagraphStyle(
        'ModeText',
        parent=styles['Normal'],
        fontSize=11,
        spaceAfter=7,
        fontName='Helvetica-Bold',
        textColor=colors.black
    )))
    
    info_data = [
        ['Duration', ':', duration_str, 'IMSI Detected', ':', str(imsi_detected)],
        ['Start Date', ':', start_date_str, 'Alert', ':', str(alert_count)],
        ['End Date', ':', end_date_str, '', '', '']
    ]
    
    info_table = Table(info_data, colWidths=[1.0*inch, 0.2*inch, 3.0*inch, 1.2*inch, 0.2*inch, 1.9*inch])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 0.3*inch))
    
    crawling_table = Table(table_data, colWidths=[0.5*inch, 1.5*inch, 1.5*inch, 0.8*inch, 1.2*inch, 2.0*inch])
    
    crawling_table.setStyle(TableStyle([
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#616161')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        
        ('ALIGN', (0, 1), (-1, -1), 'CENTER'), 
        ('ALIGN', (1, 1), (1, -1), 'LEFT'), 
        ('ALIGN', (5, 1), (5, -1), 'LEFT'), 
        
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        
        # Borders: Horizontal Only - Remove GRID, use LINEBELOW
        ('LINEBELOW', (0, 0), (-1, -1), 0.5, colors.lightgrey),
        
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9f9f9')])
    ]))
    
    elements.append(crawling_table)
    
    doc.build(elements, onFirstPage=draw_header_footer, onLaterPages=draw_header_footer)
    pdf_buffer.seek(0)
    return pdf_buffer.getvalue()


def generate_excel(db: Session, campaign_id: int) -> bytes:
    """Generate Excel export dengan openpyxl"""
    data = get_campaign_with_crawling(db, campaign_id)
    
    if data["status"] == "error":
        return None
    
    campaign = data["campaign"]
    crawlings = data["crawlings"]
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Crawling Report"
    
    # Set column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 20

    info_font = Font(bold=True, size=11, color="2c3e50")
    info_fill = PatternFill(start_color="ecf0f1", end_color="ecf0f1", fill_type="solid")
    
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="123467", end_color="123467", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    border = Border(
        left=Side(style='thin', color='bdc3c7'),
        right=Side(style='thin', color='bdc3c7'),
        top=Side(style='thin', color='bdc3c7'),
        bottom=Side(style='thin', color='bdc3c7')
    )

    # Campaign info section
    row = 1
    
    # Row labels
    labels = [
        ("Campaign Name:", campaign.name),
        ("Target IMSI:", campaign.imsi),
        ("Provider:", campaign.provider),
        ("Created At:", campaign.created_at.strftime("%Y-%m-%d %H:%M:%S") if campaign.created_at and hasattr(campaign.created_at, 'strftime') else str(campaign.created_at) if campaign.created_at else "-"),
    ]
    
    for label, value in labels:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = info_font
        ws[f'A{row}'].fill = info_fill
        ws[f'A{row}'].border = border
        
        ws[f'B{row}'] = value
        ws[f'B{row}'].border = border
        ws[f'B{row}'].alignment = Alignment(horizontal="left", vertical="center")
        
        row += 1
    
    # Crawling data table
    row += 1
    
    headers = ["IMSI", "Channel", "Provider", "Lat/Long", "UL RSSI", "Timestamp"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    ws.row_dimensions[row].height = 20
    
    row += 1
    
    # Data rows
    for crawl in crawlings:
        # Handle timestamp - can be string or datetime
        if hasattr(crawl, 'timestamp') and crawl.timestamp:
            if isinstance(crawl.timestamp, str):
                timestamp_str = crawl.timestamp
            else:
                timestamp_str = crawl.timestamp.strftime("%Y-%m-%d %H:%M:%S")
        else:
            timestamp_str = "-"
        
        ulrssi_str = str(crawl.ulRssi) if hasattr(crawl, 'ulRssi') and crawl.ulRssi is not None else "-"
        
        data_row = [
            crawl.imsi,
            "CH-01",  # TODO: update nanti dari data model
            "Telkomsel",  # TODO: update nanti dari data model
            "-6.2088,106.8456",  # TODO: update nanti dari data model
            ulrssi_str,
            timestamp_str
        ]
        
        for col_num, value in enumerate(data_row, 1):
            cell = ws.cell(row=row, column=col_num)
            cell.value = value
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = border
            
            # Alternate row colors
            if row % 2 == 0:
                cell.fill = PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type="solid")
        
        row += 1
    
    # Save to bytes
    excel_bytes = io.BytesIO()
    wb.save(excel_bytes)
    excel_bytes.seek(0)
    return excel_bytes.getvalue()
